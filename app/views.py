import os
from collections import defaultdict
from glob import glob
from urllib.parse import quote

import aiohttp_jinja2
import openpyxl
from aiohttp import web

from app.tpdf import FILES, TPdf


class ResponseFile(web.Response):
    def __new__(cls, file_name, file_body):
        headers = {
            "Content-Type": "application/pdf; charset='utf-8'",
            "Content-Disposition": "inline; filename*=UTF-8''{}".format(
                quote(file_name, encoding="utf-8"))
        }
        return web.Response(body=file_body, headers=headers)


@aiohttp_jinja2.template("index.html")
async def index(request):
    data = {}
    for f in os.listdir(FILES):
        sub_dir = os.path.join(FILES, f)
        if os.path.isfile(sub_dir):
            continue
        xlsx_path = os.path.join(sub_dir, "data.xlsx")
        data[f] = {
            "has_data_xlsx": os.path.isfile(xlsx_path),
        }
        # если нет файла настроек полей и есть файл с данными, то формируем файл настроек
        if not os.path.isfile(os.path.join(sub_dir, "fields.json")) and data[f]["has_data_xlsx"]:
            new_pos = defaultdict(list)
            wb_obj = openpyxl.load_workbook(xlsx_path)
            sheet_obj = wb_obj.active
            y_by_page_num = dict()
            for i in range(1, sheet_obj.max_column + 1):
                page_num = str(int(sheet_obj.cell(row=1, column=i).value) - 1)
                field_name = sheet_obj.cell(row=2, column=i).value
                y_by_page_num.setdefault(page_num, 720)
                new_pos[page_num].append([50, y_by_page_num[page_num], field_name, "Times New Roman", 10, 400])
                y_by_page_num[page_num] -= 30
            TPdf.save_fields_to_file(dict(new_pos, **{"dir_name": f}))
    return {"data": data}


@aiohttp_jinja2.template("positioning.html")
async def positioning(request):
    tpdf = TPdf()
    # дефолтные параметры
    in_data = {
        "dir_name": "ClearPage",
        "page_num": "1",
    }
    in_data.update(dict(request.query))
    fields = tpdf.load_fields_from_file(name=in_data["dir_name"], to_front=True)
    fonts = [os.path.basename(filename)[:-4] for filename in glob(os.path.join(tpdf.FONTS, "*.ttf"))]
    in_data.update({"fields": fields, "fonts": fonts})
    return in_data


async def save_form_fields(request):
    rq = await request.json()
    return TPdf.save_fields_to_file(rq["pos"])


async def get_file(request):
    dir_name = request.query["dir_name"]
    tpdf = TPdf()
    file = tpdf.get_pdf(dir_name, b64="False", fill_x=True)
    return ResponseFile(dir_name, file)


async def get_file_with_data(request):
    dir_name = request.query["dir_name"]
    tpdf = TPdf()
    file = tpdf.get_pdf_with_data(dir_name, b64="False")
    return ResponseFile(dir_name, file)


async def example(request):
    # набор данных для генерации комплекта документов
    data = {
        "last_name": "Иванова",
        "first_name": "Мария",
        "middle_name": "Иванова",
        "gender": "Ж",
        "birth_date": "01.01.2000",
        "birth_place": "г.Москва",
        "registration": "г.Москва, ул. Полковника Исаева, дом 17, кв 43",
        "1_work": "Радистка 3 категории, в/ч 89031",
        "2_work": "Радистка 1 категории, в/ч 17043",
        "3_work": "Командир отделения радистов, в/ч 17043 главного управления разведки комитета государственной безопасности республики Беларусь.",
    }

    # перечень документов в комплекте
    complete = [
        ("ZayavlenieNaZagranpasport", 1),
        ("ClearPage", 1),
        ("ZayavlenieNaZagranpasport", 1),
    ]

    # загружаем данные в основной класс и получаем комплект документов в pdf
    tpdf = TPdf()
    # можно сгенерировать один файл или комплект документов
    # file = tpdf.get_pdf("ZayavlenieNaZagranpasport ", b64="False")
    file = tpdf.get_complete(complete, data, b64="False")

    return ResponseFile("ZayavlenieNaZagranpasport.pdf", file)
